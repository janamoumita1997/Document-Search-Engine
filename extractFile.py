# import parser object from tike
from tika import parser 
import numpy as np
import pathlib
import PyPDF2
import re
import glob
import os
import pandas as pd
from openpyxl import load_workbook
import xlrd
import pymongo
import json
from pdf_content_extractor import *
from gensim.parsing.preprocessing import remove_stopwords
#from pptx import Presentation


class ExtractFile:
    def __init__(self, fileName,fileDirectory):
        #self.fileDirectory = "/home/mentech/Moumita/metadata_extraction_poc/source_file/"
        self.fileDirectory = fileDirectory 
        self.textFilePath = "/home/mentech/Moumita/metadata_extraction_poc/extractedFiles/"
        self.fileName = fileName
    
    def extractPDF(self):
        try:
            file = self.fileDirectory + self.fileName
            extension=self.fileName.split(".")[-1]
            #print(f"extension : {extension}")
            FileName = self.fileName.split(".")[0].replace(" ", "_") +'.'+extension
            print(f"textFileName :  {FileName}")
            parsed_pdf = parser.from_file(file)
            metadata=parsed_pdf.get("metadata")

            main_content,stop_content=content_extractor(file).content()

            result={"filename": FileName,"main_content":main_content,"stop_content":stop_content, "metadata":json.dumps(metadata),"extension":extension} 
            return result
        except Exception as e:
            print(e)
    
    def extractDOC(self):
        try:
            writeText = ''
            file = self.fileDirectory + self.fileName
            extension= self.fileName.split(".")[-1]
            textFileName = self.fileName.split(".")[0].replace(" ", "_") + '.'+extension
            parsed = parser.from_file(file)
            #print(f"parsed : {parsed}")
            metadata=parsed.get("metadata")
            #print(f"metadata : {metadata}")
            totaltext = parsed["content"]
            #print(f"totaltext : {totaltext}")
            splitline = totaltext.splitlines()
            #print(f"splitline: {splitline}")
            for textpath in splitline:
                #print("textpath : ",textpath)
                if (textpath != ''):
                    writeText+='  '+ textpath

            filtered_sentence = remove_stopwords(writeText)

            new_matadata= json.dumps(metadata)
            #print(f"metadata          >>>>>{new_matadata}")

            result={"filename": textFileName, "metadata":json.dumps(metadata), "main_content":writeText,"stop_content":filtered_sentence,"extension":extension}
            print("docx")
            return result
        except Exception as e:
            print(e)

        
    def get_sheetnames_xlsx(self, file):
        filepath = self.fileDirectory + self.fileName
        filename, file_extension = os.path.splitext(filepath)
        if file_extension != '.xls':
            wb = load_workbook(filepath, read_only=True, keep_links=False)
            return wb.sheetnames
        else:
            xls = xlrd.open_workbook(filepath, on_demand=True)
            return xls.sheet_names()
    def date_string(self,data):
        value=None
        if data != None:
            value=data.strftime("%B %d, %Y")
        return value

    def extractExcel(self):
        try:
            file = self.fileDirectory + self.fileName
            #print(f"file  : {file}")
            wb = load_workbook(file)
            x=wb.properties
            excel_metadata={"creator":x.creator,"title":x.title,"description":x.description,"subject":x.subject,"identifier":x.identifier,"language":x.language,"created":self.date_string(x.created) ,"modified":self.date_string(x.modified),"lastModifiedBy":x.lastModifiedBy,"category":x.category,"contentStatus":x.contentStatus,"version":x.version,"revision":x.revision,"keywords":x.keywords,"lastPrinted":self.date_string(x.lastPrinted)}
            #print(excel_metadata)
            #print(f"excel_metadata : {excel_metadata}")
            head, fileName = os.path.split(file)
            #print(f"head :{head}, filename :{fileName}")
            extension=fileName.split(".")[-1]
            textFileName = fileName.split(".")[0].replace(" ", "_") +'.'+ extension
            content=''
            for sheet in self.get_sheetnames_xlsx(file):
                df = pd.read_excel(file, sheet_name=sheet)
                df.fillna("")
                df = df.replace(np.nan, "")
                df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
                df = df.applymap(str)
                shape=df.shape
                #print(df)
                columns = list(df.columns)
                
                for i, d in df.iterrows():
                    #print("d   >>>>>>",d)
                    abc=' '.join(d[i] for i in range (shape[1]))
                    content += "  "+abc

            filtered_sentence = remove_stopwords(content)

                    

            result={"filename": textFileName,"metadata":json.dumps(excel_metadata), "main_content":content,"stop_content":filtered_sentence,"extension":extension}   #"metadata":excel_metadata,
            print("xlsx")
            return result
        except Exception as e:
            print(e)
    
    # def extractPPT(self):
    #     try:
    #         file = self.fileDirectory + self.fileName
    #         head, fileName = os.path.split(file)
    #         textFileName = fileName.split(".")[0].replace(" ", "_") + '.txt'
    #         parsed_pdf = Presentation(file)
    #         with open(self.textFilePath + textFileName, "w") as f:
    #             for slide in parsed_pdf.slides: 
    #                 for shape in slide.shapes: 
    #                     if not shape.has_text_frame: 
    #                         continue 
    #                     for paragraph in shape.text_frame.paragraphs: 
    #                         for run in paragraph.runs: 
    #                             f.write(run.text)
    #                         f.write("\n")
            
    #             f.close()           
    #         return textFileName
    #     except Exception as e:
    #         print(e)

    def getExtractedTextFile(self):
        try:
            if (self.fileName.endswith(".docx")) or (self.fileName.endswith(".doc")):
                return self.extractDOC()
            elif self.fileName.endswith(".pdf"):
                return self.extractPDF()
            elif (self.fileName.endswith(".xlsx")) or (self.fileName.endswith(".xls")):
                return self.extractExcel()
            # elif self.fileName.endswith(".pptx"):
            #     return self.extractPPT
            else:
                return None
        except Exception as e:
            print(e)
            return None



# fileDirectory='/home/mentech/Moumita/metadata_extraction_poc/uploadedFiles/'
# extractfile=ExtractFile("Accountant JD General 29 Jan 2020 2020-01-29.xlsx",fileDirectory)
# rocords=extractfile.getExtractedTextFile()
# print(rocords.get("metadata"))







  
